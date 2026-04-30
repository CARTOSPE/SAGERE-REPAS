#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SAGERE — Application de commandes repas traiteur
Version finale — fonctionne avec fichiers JSON locaux OU Google Sheets
"""

import streamlit as st
import json, os, re, io
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(
    page_title="SAGERE · Commandes Repas",
    page_icon="🍽",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Chemins données locales ──────────────────────────────────────────────────
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.path.join(BASE_DIR, "data")
DATA_FILE     = os.path.join(DATA_DIR, "commandes.json")
MENUS_FILE    = os.path.join(DATA_DIR, "menus.json")
SALARIES_FILE = os.path.join(DATA_DIR, "salaries.json")
CARTE_FILE    = os.path.join(DATA_DIR, "carte.json")
os.makedirs(DATA_DIR, exist_ok=True)

# ─── Constantes ───────────────────────────────────────────────────────────────
JOURS    = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
CAT_MENU = ["Entrées", "Plats garnis", "Accompagnements", "Produits laitiers", "Desserts"]
CAT_COLORS = {
    "Entrées":"#8B6BBF", "Plats garnis":"#3E7EC4",
    "Accompagnements":"#2EA86A", "Produits laitiers":"#D4902A",
    "Desserts":"#C4546A", "Carte du jour":"#3AACAC",
}
CAT_ICONS = {
    "Entrées":"🥗 ", "Plats garnis":"🍖 ",
    "Accompagnements":"🥦 ", "Produits laitiers":"🧀 ", "Desserts":"🍮 ",
}
DEFAULT_SALARIES = ["GHEYSENS Eric","CAMPION Pascal","CHARPENTIER Franck","PEREIRA Serge"]
DEFAULT_CARTE = {
    "Entrées":           ["Tomate et dosette de vinaigrette","Salade verte","Œuf dur mayonnaise"],
    "Plats garnis":      ["Filet de poulet","Jambon blanc","Pané de blé, tomate et mozzarella",
                          "Pavé de colin mariné huile d'olive et citron vert","Steak haché cuit à cœur"],
    "Accompagnements":   ["Pommes vapeur","Frites au four","Pâtes","Haricots verts"],
    "Produits laitiers": [],
    "Desserts":          ["Crème dessert chocolat","Purée de pommes fraises","Tarte aux pommes"],
}

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background:#181B2E !important; }
[data-testid="stSidebar"]          { background:#22263D !important; }
[data-testid="stSidebar"] > div    { padding-top:1rem; }
.block-container { padding-top:1.4rem; padding-bottom:2rem; }
div[data-testid="stCheckbox"] label p { color:#E8EAF6 !important; font-size:0.95rem; }
.sidebar-label { font-size:0.70rem;font-weight:700;color:#555A82;letter-spacing:0.12em;margin:14px 0 3px 0; }
.recap-ok  { color:#3DBE6E; font-weight:700; }
.recap-non { color:#555A82; }
div[data-testid="stVerticalBlock"] > div { background:transparent !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# COUCHE PERSISTANCE — JSON local + Google Sheets optionnel
# ══════════════════════════════════════════════════════════════════════════════

def _use_gsheets():
    """Retourne True si Google Sheets est configuré et utilisable."""
    try:
        sid = st.secrets["gsheet"]["spreadsheet_id"]
        return bool(sid) and sid != "COLLER_ID_DU_GOOGLE_SHEET_ICI"
    except Exception:
        return False

# ── JSON local ────────────────────────────────────────────────────────────────
def _load_json(path, default):
    if os.path.exists(path):
        try:
            with open(path,"r",encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return default

def _save_json(path, data):
    with open(path,"w",encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── Google Sheets ─────────────────────────────────────────────────────────────
@st.cache_resource
def _gs_client():
    import gspread
    from google.oauth2.service_account import Credentials
    sa = {k:v for k,v in st.secrets["gcp_service_account"].items()}
    creds = Credentials.from_service_account_info(
        sa, scopes=["https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive"])
    return gspread.authorize(creds)

def _gs_sheet(tab):
    sh = _gs_client().open_by_key(st.secrets["gsheet"]["spreadsheet_id"])
    try:
        return sh.worksheet(tab)
    except Exception:
        return sh.add_worksheet(title=tab, rows=500, cols=3)

def _gs_load(tab):
    ws   = _gs_sheet(tab)
    rows = ws.get_all_values()
    out  = {}
    for row in rows:
        if len(row) >= 2 and row[0].strip():
            try:    out[row[0]] = json.loads(row[1])
            except: pass
    return out

def _gs_save(tab, key, value):
    ws   = _gs_sheet(tab)
    rows = ws.get_all_values()
    val  = json.dumps(value, ensure_ascii=False)
    for i, row in enumerate(rows, 1):
        if row and row[0] == key:
            ws.update(f"B{i}", [[val]]); return
    ws.append_row([key, val])

# ── API unifiée ───────────────────────────────────────────────────────────────
def _gs_error(e):
    """Stocke l'erreur Sheets pour affichage unique dans init_state."""
    msg = f"Google Sheets inaccessible (erreur {type(e).__name__}: {e}) — données locales utilisées."
    if "_gs_error" not in st.session_state:
        st.session_state._gs_error = msg

def load_menus():
    if _use_gsheets():
        try:
            data = _gs_load("menus")
            if data: return data
        except Exception as e:
            _gs_error(e)
    data = _load_json(MENUS_FILE, {})
    if not data:
        wk   = week_key()
        data = {wk:{"semaine":wk,"periode":"","jours":{j:{c:[] for c in CAT_MENU} for j in JOURS}}}
    return data

def save_menu(wk, menu):
    _save_json(MENUS_FILE, {**load_menus_local(), wk: menu})
    if _use_gsheets():
        try:    _gs_save("menus", wk, menu)
        except Exception as e: st.error(f"⚠️ Sheets (écriture) : {e}")

def load_menus_local():
    return _load_json(MENUS_FILE, {})

def load_commandes():
    if _use_gsheets():
        try:
            data = _gs_load("commandes")
            if data: return data
        except Exception as e:
            _gs_error(e)
    return _load_json(DATA_FILE, {})

def save_commandes_wk(wk, cmds):
    all_cmds = _load_json(DATA_FILE, {})
    all_cmds[wk] = cmds
    _save_json(DATA_FILE, all_cmds)
    if _use_gsheets():
        try:    _gs_save("commandes", wk, cmds)
        except Exception as e: st.error(f"⚠️ Sheets (écriture) : {e}")

def load_salaries():
    if _use_gsheets():
        try:
            data = _gs_load("config")
            if "salaries" in data: return data["salaries"]
        except Exception as e:
            _gs_error(e)
    return _load_json(SALARIES_FILE, DEFAULT_SALARIES)

def save_salaries(s):
    _save_json(SALARIES_FILE, s)
    if _use_gsheets():
        try:    _gs_save("config","salaries",s)
        except Exception as e: st.error(f"⚠️ Sheets (écriture) : {e}")

def load_carte():
    if _use_gsheets():
        try:
            data = _gs_load("config")
            if "carte" in data: return data["carte"]
        except Exception as e:
            _gs_error(e)
    return _load_json(CARTE_FILE, DEFAULT_CARTE)

def save_carte(c):
    _save_json(CARTE_FILE, c)
    if _use_gsheets():
        try:    _gs_save("config","carte",c)
        except Exception as e: st.error(f"⚠️ Sheets : {e}")

# ─── Utilitaires semaine ──────────────────────────────────────────────────────
def week_key(d=None):
    d = d or date.today()
    iso = d.isocalendar()
    return f"{iso[0]}-S{iso[1]:02d}"

def week_label(key):
    try:
        yr, sw = key.split("-S"); yr, sw = int(yr), int(sw)
        monday = date.fromisocalendar(yr, sw, 1)
        friday = monday + timedelta(days=4)
        mois = ["","jan.","fév.","mars","avr.","mai","juin",
                "juil.","août","sept.","oct.","nov.","déc."]
        return f"S{sw:02d} · {monday.day} {mois[monday.month]} – {friday.day} {mois[friday.month]} {yr}"
    except Exception:
        return key

def weeks_list(menus):
    keys = sorted(menus.keys(), reverse=True)
    return keys if keys else [week_key()]

# ─── Import menu traiteur ─────────────────────────────────────────────────────
def parse_traiteur_html(raw):
    from bs4 import BeautifulSoup
    content = None
    for enc in ("utf-8","iso-8859-1","cp1252"):
        try: content = raw.decode(enc); break
        except: pass
    if not content: raise ValueError("Impossible de décoder le fichier.")
    soup    = BeautifulSoup(content,"html.parser")
    periode_raw = ""
    p = soup.find("p", class_="block_date")
    if p: periode_raw = p.get_text(strip=True)

    # Extraire la clé semaine depuis le texte brut (ex: "2026-S23-Menu...")
    m  = re.search(r'(\d{4})-S(\d{2})', periode_raw)
    sk = f"{m.group(1)}-S{m.group(2)}" if m else week_key()

    # Nettoyer la période via la fonction centrale
    periode_clean = clean_periode(periode_raw, sk)

    tables     = soup.find_all("table", class_="table_recette")
    jours_data = {j:{c:[] for c in CAT_MENU} for j in JOURS}
    if len(tables) >= 25:
        for ci,cat in enumerate(CAT_MENU):
            for ji,jour in enumerate(JOURS):
                t = tables[ci*5+ji]
                jours_data[jour][cat] = [
                    tr.get_text(strip=True) for tr in t.find_all("tr")
                    if tr.get_text(strip=True)]
    return {"semaine":sk, "periode":periode_clean, "jours":jours_data}

# ─── UI helpers ───────────────────────────────────────────────────────────────
def clean_periode(raw, wk=None):
    """
    Nettoie la période : 'YYYY-SNN-Menu... - Du 04 mai 2026 au 08 mai 2026 -'
    → 'Du 04 mai 2026 au 08 mai 2026'
    Si vide ou illisible → week_label(wk)
    """
    if not raw or not raw.strip():
        return week_label(wk) if wk else ""
    raw = raw.strip()
    # Chercher "Du XX mois YYYY au XX mois YYYY"
    m = re.search(r'(Du\s+\d+\s+\w+\s+\d{4}\s+au\s+\d+\s+\w+\s+\d{4})', raw, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # Fallback : supprimer le préfixe YYYY-SNN et les mots parasites
    cleaned = re.sub(r'^\d{4}-S\d{2}[-_\s]*', '', raw)
    cleaned = re.sub(r'^[\w\s]+(jours?|days?)\s*[-–]\s*', '', cleaned, flags=re.IGNORECASE)
    cleaned = cleaned.strip(' -–')
    # Si encore illisible, utiliser le label calculé
    if len(cleaned) > 60 or re.search(r'\d{4}-S\d{2}', cleaned):
        return week_label(wk) if wk else raw
    return cleaned or (week_label(wk) if wk else raw)

def _periode_coherente(periode, wk):
    """Retourne True si la période correspond au lundi de la semaine wk."""
    try:
        yr, sw = wk.split("-S"); yr, sw = int(yr), int(sw)
        monday = date.fromisocalendar(yr, sw, 1)
        mois_fr = {"janvier":1,"février":2,"mars":3,"avril":4,"mai":5,
                   "juin":6,"juillet":7,"août":8,"septembre":9,
                   "octobre":10,"novembre":11,"décembre":12}
        m = re.search(r'(\d+)\s+(\w+)\s+(\d{4})', periode)
        if m:
            jour_n = int(m.group(1))
            mois_n = mois_fr.get(m.group(2).lower(), 0)
            annee  = int(m.group(3))
            if mois_n and date(annee, mois_n, jour_n) == monday:
                return True
        return False
    except Exception:
        return False

def cat_header(color, text, icon=""):
    """Bandeau coloré auto-fermé — sans div ouvert."""
    return (
        f'<div style="background:{color};padding:7px 16px;border-radius:8px 8px 0 0;'
        f'font-weight:700;font-size:0.83rem;letter-spacing:0.07em;color:#fff;'
        f'margin-top:14px;margin-bottom:2px;">{icon}{text}</div>'
    )

def inline_label(color, icon, text):
    return (f'<div style="background:{color};padding:6px 14px;border-radius:6px;'
            f'font-weight:700;font-size:0.82rem;color:#fff;margin:10px 0 4px 0;">'
            f'{icon}{text}</div>')

# ─── Exports Excel ────────────────────────────────────────────────────────────
def build_export_traiteur(menu, commandes, salaries, carte, periode, wk):
    if not periode: periode = week_label(wk)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Commande traiteur"
    def fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    thin = Border(**{s:Side(style="thin",color="444870") for s in ("left","right","top","bottom")})
    ws.merge_cells("A1:H1"); ws["A1"] = f"SAGERE — Bon de commande traiteur  |  {periode}"
    ws["A1"].font=Font(name="Calibri",bold=True,size=14,color="FFFFFF"); ws["A1"].fill=fill("1E2240")
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=32
    ws.merge_cells("A2:H2"); ws["A2"]=f"Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A2"].font=Font(name="Calibri",size=9,color="888EC0"); ws["A2"].fill=fill("1E2240")
    ws["A2"].alignment=Alignment(horizontal="center"); ws.row_dimensions[2].height=18
    for c,h in enumerate(["Catégorie","Plat / Article"]+JOURS+["TOTAL SEMAINE"],1):
        cell=ws.cell(row=3,column=c,value=h)
        cell.font=Font(name="Calibri",bold=True,size=10,color="E8EAF6"); cell.fill=fill("2B3270")
        cell.alignment=Alignment(horizontal="center" if c>2 else "left"); cell.border=thin
    ws.row_dimensions[3].height=22
    cat_hex={"Entrées":"3A2060","Plats garnis":"1A3A68","Accompagnements":"0D4A2A",
             "Produits laitiers":"5A3A08","Desserts":"5A1A28","Carte du jour":"0A3A3A"}
    cat_fg ={"Entrées":"D8C0F8","Plats garnis":"C0D8F8","Accompagnements":"B0F0D0",
             "Produits laitiers":"F8E0A0","Desserts":"F8C0CC","Carte du jour":"A0E8E8"}
    row=[4]; grand_total=[0]
    def write_block(cat_key,label,items_fn,cmd_cat):
        all_items,seen=[],set()
        for jour in JOURS:
            for it in items_fn(jour):
                if it and it not in seen: all_items.append(it); seen.add(it)
        if not all_items: return
        ws.merge_cells(start_row=row[0],start_column=1,end_row=row[0],end_column=8)
        cell=ws.cell(row=row[0],column=1,value=f"  ▸  {label.upper()}")
        cell.font=Font(name="Calibri",bold=True,size=10,color=cat_fg.get(cat_key,"E8EAF6"))
        cell.fill=fill(cat_hex.get(cat_key,"222222")); cell.alignment=Alignment(vertical="center")
        ws.row_dimensions[row[0]].height=18; row[0]+=1
        for item in all_items:
            totaux,ltotal=[],0
            for jour in JOURS:
                qty=sum(1 for sal in salaries if item in commandes.get(sal,{}).get(jour,{}).get(cmd_cat,[]))
                totaux.append(qty); ltotal+=qty
            grand_total[0]+=ltotal
            ws.cell(row=row[0],column=1,value=label).font=Font(name="Calibri",size=8,color="667090")
            ws.cell(row=row[0],column=1).fill=fill("1E2240"); ws.cell(row=row[0],column=1).border=thin
            ws.cell(row=row[0],column=2,value=item).font=Font(name="Calibri",size=10,color="D8DCFF")
            ws.cell(row=row[0],column=2).fill=fill("1E2240"); ws.cell(row=row[0],column=2).border=thin
            for ji,qty in enumerate(totaux):
                c=ws.cell(row=row[0],column=3+ji,value=qty if qty else "")
                c.fill=fill("0D3020" if qty>0 else "1E2240")
                c.font=Font(name="Calibri",bold=(qty>0),size=10,color="60E890" if qty>0 else "444870")
                c.alignment=Alignment(horizontal="center"); c.border=thin
            ct=ws.cell(row=row[0],column=8,value=ltotal if ltotal else "")
            ct.fill=fill("101428"); ct.border=thin
            ct.font=Font(name="Calibri",bold=True,size=10,color="FFD060" if ltotal>0 else "444870")
            ct.alignment=Alignment(horizontal="center"); ws.row_dimensions[row[0]].height=16; row[0]+=1
    for cat in CAT_MENU:
        write_block(cat,cat,lambda jour,c=cat:menu.get("jours",{}).get(jour,{}).get(c,[]),cat)
    if any(carte.get(s) for s in CAT_MENU):
        ws.merge_cells(start_row=row[0],start_column=1,end_row=row[0],end_column=8)
        sep=ws.cell(row=row[0],column=1,value="  ━━━  CARTE DU JOUR (permanente)  ━━━")
        sep.font=Font(name="Calibri",bold=True,size=11,color="A0E8E8"); sep.fill=fill("0A3A3A")
        sep.alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[row[0]].height=20; row[0]+=1
        for sub in CAT_MENU:
            items=carte.get(sub,[])
            if items: write_block("Carte du jour",f"Carte · {sub}",lambda jour,it=items:it,f"Carte · {sub}")
    row[0]+=1
    ws.merge_cells(start_row=row[0],start_column=1,end_row=row[0],end_column=2)
    ws.cell(row=row[0],column=1,value="TOTAL JOURNALIER").font=Font(bold=True,color="FFD060",size=10)
    ws.cell(row=row[0],column=1).fill=fill("101428")
    for ji,jour in enumerate(JOURS):
        tj=sum(len(commandes.get(sal,{}).get(jour,{}).get(cat,[]))
               for sal in salaries for cat in CAT_MENU+[f"Carte · {s}" for s in CAT_MENU])
        c=ws.cell(row=row[0],column=3+ji,value=tj)
        c.font=Font(bold=True,size=11,color="FFD060"); c.fill=fill("101428")
        c.alignment=Alignment(horizontal="center"); c.border=thin
    ws.cell(row=row[0],column=8,value=grand_total[0]).font=Font(bold=True,size=12,color="FFD060")
    ws.cell(row=row[0],column=8).fill=fill("101428")
    ws.cell(row=row[0],column=8).alignment=Alignment(horizontal="center"); ws.row_dimensions[row[0]].height=24
    ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=44
    for j in range(5): ws.column_dimensions[openpyxl.utils.get_column_letter(3+j)].width=12
    ws.column_dimensions["H"].width=14; ws.freeze_panes="C4"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def build_export_interne(menu, commandes, salaries, carte, periode, wk):
    if not periode: periode = week_label(wk)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Récapitulatif"
    def fill(h): return PatternFill("solid",fgColor=h.lstrip("#"))
    thin=Border(**{s:Side(style="thin",color="444870") for s in ("left","right","top","bottom")})
    nb=2+len(JOURS)*len(salaries)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=nb)
    ws["A1"]=f"SAGERE — Récapitulatif interne  |  {periode}"
    ws["A1"].font=Font(bold=True,size=14,color="FFFFFF"); ws["A1"].fill=fill("1E2240")
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30
    col=3
    for jour in JOURS:
        ws.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+len(salaries)-1)
        c=ws.cell(row=2,column=col,value=jour.upper())
        c.font=Font(bold=True,size=10,color="FFFFFF"); c.fill=fill("2B3270")
        c.alignment=Alignment(horizontal="center"); col+=len(salaries)
    ws.cell(row=3,column=1,value="Catégorie").font=Font(bold=True,color="FFFFFF")
    ws.cell(row=3,column=2,value="Article").font=Font(bold=True,color="FFFFFF")
    ws.cell(row=3,column=1).fill=ws.cell(row=3,column=2).fill=fill("2B3270")
    col=3
    for jour in JOURS:
        for sal in salaries:
            c=ws.cell(row=3,column=col,value=sal.split()[0])
            c.font=Font(bold=True,size=8,color="FFFFFF"); c.fill=fill("363B5E")
            c.alignment=Alignment(horizontal="center",wrap_text=True); col+=1
    ws.row_dimensions[3].height=28
    cat_fg={"Entrées":"E8D5F8","Plats garnis":"CCE0F8","Accompagnements":"C8EDF0",
            "Produits laitiers":"FDE8C0","Desserts":"F8C8CF","Carte du jour":"C0ECEC"}
    row=[4]
    def write_rows(label,items,cmd_cat,fgc):
        for item in items:
            ws.cell(row=row[0],column=1,value=label).fill=fill(fgc)
            ws.cell(row=row[0],column=1).font=Font(size=8,bold=True); ws.cell(row=row[0],column=1).border=thin
            ws.cell(row=row[0],column=2,value=item).fill=fill("F8F9FF")
            ws.cell(row=row[0],column=2).font=Font(size=9); ws.cell(row=row[0],column=2).border=thin
            col=3
            for jour in JOURS:
                for sal in salaries:
                    has=item in commandes.get(sal,{}).get(jour,{}).get(cmd_cat,[])
                    c=ws.cell(row=row[0],column=col)
                    if has: c.value="✓"; c.font=Font(bold=True,color="1A7340"); c.fill=fill("D4F5E0")
                    c.alignment=Alignment(horizontal="center"); c.border=thin; col+=1
            ws.row_dimensions[row[0]].height=14; row[0]+=1
    for cat in CAT_MENU:
        items,seen=[],set()
        for jour in JOURS:
            for it in menu.get("jours",{}).get(jour,{}).get(cat,[]):
                if it and it not in seen: items.append(it); seen.add(it)
        write_rows(cat,items,cat,cat_fg.get(cat,"EEEEEE"))
    for sub in CAT_MENU:
        items=carte.get(sub,[])
        if items: write_rows(f"Carte · {sub}",items,f"Carte · {sub}",cat_fg["Carte du jour"])
    ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=40
    for i in range(3,3+len(JOURS)*len(salaries)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=9
    ws.freeze_panes="C4"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ══════════════════════════════════════════════════════════════════════════════
# INIT SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
def init_state():
    if "loaded" not in st.session_state:
        with st.spinner("⏳ Chargement…"):
            st.session_state.menus     = load_menus()
            st.session_state.commandes = load_commandes()
            st.session_state.salaries  = load_salaries()
            st.session_state.carte     = load_carte()
            st.session_state.loaded    = True
            # Mode réel : Sheets si configuré ET sans erreur 404
            gs_ok = _use_gsheets() and "_gs_error" not in st.session_state
            st.session_state._storage_mode = "Google Sheets ☁️" if gs_ok else "Fichiers locaux 💾"

    # Afficher l'erreur Sheets une seule fois, clairement
    if "_gs_error" in st.session_state:
        err = st.session_state.pop("_gs_error")
        if "404" in str(err):
            st.warning(
                "⚠️ **Google Sheets introuvable (erreur 404)**\n\n"
                "L'ID du Sheet dans `secrets.toml` est incorrect ou le Sheet n'est pas partagé.\n\n"
                "**À vérifier :**\n"
                "1. Ouvrez votre Google Sheet → copiez l'ID depuis l'URL (entre `/d/` et `/edit`)\n"
                "2. Collez-le dans `secrets.toml` → `spreadsheet_id = \"VOTRE_ID\"`\n"
                "3. Vérifiez que le Sheet est partagé avec `sagere-sheets@sagere.iam.gserviceaccount.com`\n\n"
                "*En attendant, les données sont sauvegardées localement.*"
            )
        else:
            st.warning(f"⚠️ Google Sheets inaccessible — données locales utilisées.\n\n`{err}`")

    if "page"     not in st.session_state: st.session_state.page     = "commande"
    if "week_sel" not in st.session_state: st.session_state.week_sel = week_key()
    if "salarie"  not in st.session_state:
        st.session_state.salarie = st.session_state.salaries[0] if st.session_state.salaries else ""
    if "jour"     not in st.session_state:
        st.session_state.jour = JOURS[min(date.today().weekday(), 4)]
    wks = weeks_list(st.session_state.menus)
    if st.session_state.week_sel not in wks:
        st.session_state.week_sel = wks[0]

init_state()

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🍽 SAGERE")
    st.markdown("*Commandes repas traiteur*")
    mode = st.session_state.get("_storage_mode","")
    if mode:
        st.caption(f"Stockage : {mode}")
    st.divider()

    for key, label in {
        "commande":"🧾 Passer commande", "menu":"⚙ Saisir le menu",
        "carte":"🗂 Carte permanente",   "salaries":"👥 Salariés",
        "admin":"📊 Exports & Admin"
    }.items():
        if st.button(label, key=f"nav_{key}",
                     type="primary" if st.session_state.page==key else "secondary",
                     use_container_width=True):
            st.session_state.page=key; st.rerun()

    st.divider()

    # Semaine
    st.markdown('<p class="sidebar-label">SEMAINE</p>', unsafe_allow_html=True)
    wks    = weeks_list(st.session_state.menus)
    labels = [week_label(k) for k in wks]
    if st.session_state.week_sel not in wks: st.session_state.week_sel = wks[0]
    sel = st.selectbox("Semaine", labels,
                       index=wks.index(st.session_state.week_sel),
                       key="week_select_box", label_visibility="collapsed")
    st.session_state.week_sel = wks[labels.index(sel)]


# Export déduction paie (niveau module — pas dans un elif)
# ══════════════════════════════════════════════════════════════════════════════

# Catégories qui font un "repas complet" (en plus du plat)
CAT_COMPLEMENT = {"Entrées", "Produits laitiers", "Desserts"}
CAT_PLAT       = {"Plats garnis", "Accompagnements"}

def classifier_repas(jour_cmds):
    """
    Analyse les commandes d'un salarié pour un jour.
    Retourne : 'complet' | 'plat_unique' | 'rien'
    jour_cmds = {cat: [items], ...}  (peut contenir clés 'Carte · XXX' aussi)
    """
    a_plat = False
    a_complement = False
    for cat, items in jour_cmds.items():
        if not items: continue
        # Normaliser : 'Carte · Entrées' → 'Entrées' etc.
        cat_base = cat.replace("Carte · ", "")
        if cat_base in CAT_PLAT:
            a_plat = True
        if cat_base in CAT_COMPLEMENT:
            a_complement = True
    if not a_plat and not a_complement:
        return "rien"
    if a_plat and a_complement:
        return "complet"
    if a_plat and not a_complement:
        return "plat_unique"
    # Seulement entrée/dessert sans plat → complet quand même
    if a_complement:
        return "complet"
    return "rien"

def build_export_paie(commandes_multi, salaries, semaines):
    """
    Export déduction paie multi-semaines.
    commandes_multi = {wk: {sal: {jour: {cat: [items]}}}}
    semaines = [(wk, label, periode), ...]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Déduction paie"

    def fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    thin = Border(**{s: Side(style="thin", color="444870")
                     for s in ("left","right","top","bottom")})
    thick_left = Border(left=Side(style="medium", color="5B7FE8"),
                        right=Side(style="thin", color="444870"),
                        top=Side(style="thin", color="444870"),
                        bottom=Side(style="thin", color="444870"))

    # ── Titre ──
    ws.merge_cells("A1:Z1")
    ws["A1"] = f"SAGERE — Déduction paie repas  |  Édité le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill("1E2240")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Construction du tableau : une section par semaine ──
    current_row = 2

    for wk, lbl, periode in semaines:
        wk_cmds = commandes_multi.get(wk, {})

        # En-tête semaine
        nb_cols = 1 + len(JOURS) * 2 + 2  # Salarié + (C+P)*5jours + Total C + Total P
        ws.merge_cells(start_row=current_row, start_column=1,
                        end_row=current_row, end_column=nb_cols)
        cell = ws.cell(row=current_row, column=1,
                        value=f"  {lbl}  |  {periode}")
        cell.font = Font(name="Calibri", bold=True, size=11, color="C0D8F8")
        cell.fill = fill("1A3A68")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Sous-en-têtes : Salarié | Lun C | Lun P | Mar C | ... | Total C | Total P
        ws.cell(row=current_row, column=1, value="Salarié").font = Font(bold=True, color="FFFFFF", size=9)
        ws.cell(row=current_row, column=1).fill = fill("2B3270")
        ws.cell(row=current_row, column=1).border = thin

        col = 2
        for jour in JOURS:
            # Fusionner les 2 colonnes du jour
            ws.merge_cells(start_row=current_row, start_column=col,
                            end_row=current_row, end_column=col+1)
            c = ws.cell(row=current_row, column=col, value=jour[:3].upper())
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = fill("2B3270")
            c.alignment = Alignment(horizontal="center")
            c.border = thin
            col += 2

        # Totaux
        for lbl_t, clr in [("Tot. Complet","1A4A28"),("Tot. Plat unique","3A3A1A")]:
            c = ws.cell(row=current_row, column=col, value=lbl_t)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = fill(clr)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = thin
            col += 1
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Ligne légende C/P
        ws.cell(row=current_row, column=1, value="").fill = fill("2B3270")
        ws.cell(row=current_row, column=1).border = thin
        col = 2
        for _ in JOURS:
            for lbl_cp, clr_cp in [("C","1A4A28"),("P","3A3A1A")]:
                c = ws.cell(row=current_row, column=col, value=lbl_cp)
                c.font = Font(bold=True, size=8, color="FFFFFF")
                c.fill = fill(clr_cp)
                c.alignment = Alignment(horizontal="center")
                c.border = thin
                col += 1
        for _ in range(2):
            ws.cell(row=current_row, column=col).fill = fill("2B3270")
            ws.cell(row=current_row, column=col).border = thin
            col += 1
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Lignes salariés
        for sal in salaries:
            sal_cmds = wk_cmds.get(sal, {})
            tot_c = tot_p = 0
            col = 2
            ws.cell(row=current_row, column=1, value=sal)
            ws.cell(row=current_row, column=1).font = Font(size=9, bold=True)
            ws.cell(row=current_row, column=1).border = thick_left

            for jour in JOURS:
                type_repas = classifier_repas(sal_cmds.get(jour, {}))
                est_c = (type_repas == "complet")
                est_p = (type_repas == "plat_unique")
                if est_c: tot_c += 1
                if est_p: tot_p += 1

                # Colonne Complet
                c = ws.cell(row=current_row, column=col,
                             value="✓" if est_c else "")
                c.fill = fill("0D3020" if est_c else "1E2240")
                c.font = Font(bold=True, size=10,
                               color="60E890" if est_c else "444870")
                c.alignment = Alignment(horizontal="center")
                c.border = thin
                col += 1

                # Colonne Plat unique
                c = ws.cell(row=current_row, column=col,
                             value="✓" if est_p else "")
                c.fill = fill("2A2A0A" if est_p else "1E2240")
                c.font = Font(bold=True, size=10,
                               color="E8E060" if est_p else "444870")
                c.alignment = Alignment(horizontal="center")
                c.border = thin
                col += 1

            # Totaux
            for val, clr_bg, clr_fg in [
                (tot_c, "0D3020", "60E890"),
                (tot_p, "2A2A0A", "E8E060"),
            ]:
                c = ws.cell(row=current_row, column=col,
                             value=val if val > 0 else "")
                c.font = Font(bold=True, size=11, color=clr_fg if val>0 else "444870")
                c.fill = fill(clr_bg)
                c.alignment = Alignment(horizontal="center")
                c.border = thin
                col += 1

            ws.row_dimensions[current_row].height = 16
            current_row += 1

        # Ligne totaux semaine
        ws.cell(row=current_row, column=1, value="TOTAL SEMAINE").font = Font(bold=True, size=9, color="C0D8F8")
        ws.cell(row=current_row, column=1).fill = fill("1A3A68")
        ws.cell(row=current_row, column=1).border = thin
        col = 2
        for jour in JOURS:
            tot_j_c = sum(1 for sal in salaries
                          if classifier_repas(wk_cmds.get(sal,{}).get(jour,{})) == "complet")
            tot_j_p = sum(1 for sal in salaries
                          if classifier_repas(wk_cmds.get(sal,{}).get(jour,{})) == "plat_unique")
            for val, clr in [(tot_j_c,"60E890"),(tot_j_p,"E8E060")]:
                c = ws.cell(row=current_row, column=col, value=val if val else "")
                c.font = Font(bold=True, size=10, color=clr if val else "444870")
                c.fill = fill("1A3A68")
                c.alignment = Alignment(horizontal="center")
                c.border = thin
                col += 1

        # Total général semaine (C et P)
        tot_sem_c = sum(1 for sal in salaries for jour in JOURS
                         if classifier_repas(wk_cmds.get(sal,{}).get(jour,{})) == "complet")
        tot_sem_p = sum(1 for sal in salaries for jour in JOURS
                         if classifier_repas(wk_cmds.get(sal,{}).get(jour,{})) == "plat_unique")
        for val, clr in [(tot_sem_c,"60E890"),(tot_sem_p,"E8E060")]:
            c = ws.cell(row=current_row, column=col, value=val if val else "")
            c.font = Font(bold=True, size=11, color=clr if val else "444870")
            c.fill = fill("1A3A68")
            c.alignment = Alignment(horizontal="center")
            c.border = thin
            col += 1
        ws.row_dimensions[current_row].height = 18
        current_row += 2  # Ligne vide entre semaines

    # ── Feuille récap global multi-semaines ──
    ws2 = wb.create_sheet("Récap global")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "SAGERE — Récapitulatif global déduction paie"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = fill("1E2240")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # En-têtes
    for c, h in enumerate(["Salarié"] + [week_label(s[0]) for s in semaines] +
                            ["TOTAL Complet","TOTAL Plat unique","TOTAL Repas"], 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = fill("2B3270")
        cell.alignment = Alignment(horizontal="center" if c>1 else "left", wrap_text=True)
        cell.border = thin
    ws2.row_dimensions[2].height = 32

    for sal in salaries:
        row_data = [sal]
        grand_c = grand_p = 0
        for wk, _, _ in semaines:
            wk_cmds = commandes_multi.get(wk, {})
            sal_c = sum(1 for jour in JOURS
                         if classifier_repas(wk_cmds.get(sal,{}).get(jour,{})) == "complet")
            sal_p = sum(1 for jour in JOURS
                         if classifier_repas(wk_cmds.get(sal,{}).get(jour,{})) == "plat_unique")
            row_data.append(f"C:{sal_c}  P:{sal_p}" if (sal_c+sal_p) else "—")
            grand_c += sal_c; grand_p += sal_p
        row_data += [grand_c or "", grand_p or "", (grand_c+grand_p) or ""]
        r = ws2.max_row + 1
        for ci, val in enumerate(row_data, 1):
            c = ws2.cell(row=r, column=ci, value=val)
            c.font = Font(size=9)
            c.border = thin
            if ci == 1: c.font = Font(size=9, bold=True)
            if ci > len(semaines)+1:
                c.font = Font(bold=True, size=10,
                               color="60E890" if ci==len(semaines)+2
                               else "E8E060" if ci==len(semaines)+3 else "FFFFFF")
                c.fill = fill("0D3020" if ci==len(semaines)+2
                              else "2A2A0A" if ci==len(semaines)+3 else "1A2A4A")
            c.alignment = Alignment(horizontal="center" if ci>1 else "left")
        ws2.row_dimensions[r].height = 16

    # Légende
    r_leg = ws2.max_row + 2
    ws2.cell(row=r_leg, column=1, value="Légende :").font = Font(bold=True, size=9, color="AAAAAA")
    ws2.cell(row=r_leg+1, column=1,
             value="C = Repas complet (avec entrée et/ou produit laitier et/ou dessert)").font = Font(size=8, color="60E890")
    ws2.cell(row=r_leg+2, column=1,
             value="P = Plat unique (plat garni et/ou accompagnement uniquement, sans entrée ni dessert)").font = Font(size=8, color="E8E060")

    # Largeurs
    ws.column_dimensions["A"].width = 22
    for i in range(2, 2 + len(JOURS)*2 + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 6
    ws2.column_dimensions["A"].width = 22
    for i in range(2, 2+len(semaines)+3):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
    ws.freeze_panes = "B3"
    ws2.freeze_panes = "B3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    if st.session_state.page == "commande":
        # Salarié
        st.markdown('<p class="sidebar-label">SALARIÉ</p>', unsafe_allow_html=True)
        sals    = st.session_state.salaries
        sal_idx = sals.index(st.session_state.salarie) if st.session_state.salarie in sals else 0
        st.session_state.salarie = st.selectbox(
            "Salarié", sals, index=sal_idx,
            key="sal_select", label_visibility="collapsed")

        # Jour
        st.markdown('<p class="sidebar-label">JOUR</p>', unsafe_allow_html=True)
        cols = st.columns(5)
        for i, jour in enumerate(JOURS):
            with cols[i]:
                if st.button(jour[:3], key=f"jour_{jour}",
                             type="primary" if st.session_state.jour==jour else "secondary",
                             use_container_width=True):
                    st.session_state.jour=jour; st.rerun()

        # Récap
        st.divider()
        st.markdown('<p class="sidebar-label">MES COMMANDES</p>', unsafe_allow_html=True)
        sem = st.session_state.commandes.get(
            st.session_state.week_sel,{}).get(st.session_state.salarie,{})
        for jour in JOURS:
            total = sum(len(v) for v in sem.get(jour,{}).values())
            if total:
                st.markdown(f'<span class="recap-ok">✓ {jour[:3]}</span> — {total} article(s)',
                            unsafe_allow_html=True)
            else:
                st.markdown(f'<span class="recap-non">○ {jour[:3]}</span>',
                            unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE : COMMANDE
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.page == "commande":
    sal  = st.session_state.salarie
    jour = st.session_state.jour
    wk   = st.session_state.week_sel
    menu = st.session_state.menus.get(wk,{})

    c1,c2 = st.columns([2,3])
    with c1: st.markdown(f"# {jour}")
    with c2: st.markdown(f"<br><span style='color:#8890C0'>{week_label(wk)}</span>",
                         unsafe_allow_html=True)

    existing  = st.session_state.commandes.get(wk,{}).get(sal,{}).get(jour,{})
    choix     = {}
    jour_menu = menu.get("jours",{}).get(jour,{})
    carte     = st.session_state.carte
    has_menu  = any(jour_menu.get(c) for c in CAT_MENU)
    has_carte = any(carte.get(c)     for c in CAT_MENU)

    if not has_menu and not has_carte:
        st.info("📭 Aucun menu pour cette semaine. Importez ou saisissez le menu du traiteur.")
    else:
        for cat in CAT_MENU:
            items = jour_menu.get(cat,[])
            if not items: continue
            color    = CAT_COLORS[cat]
            icon     = CAT_ICONS.get(cat,"")
            selected = existing.get(cat,[])
            # Bandeau coloré
            st.markdown(cat_header(color, cat.upper(), icon), unsafe_allow_html=True)
            # Fond du bloc via container stylé
            st.markdown(
                f'<div style="background:#22263D;border:1px solid {color}55;'
                f'border-top:none;border-radius:0 0 8px 8px;padding:8px 4px 4px 4px;'
                f'margin-bottom:10px;"></div>',
                unsafe_allow_html=True)
            cols = st.columns(2)
            for i, item in enumerate(items):
                with cols[i % 2]:
                    if st.checkbox(item, value=(item in selected),
                                   key=f"cb_{wk}_{sal}_{jour}_{cat}_{i}"):
                        choix.setdefault(cat,[]).append(item)

        if has_carte:
            st.markdown(
                '<div style="background:#3AACAC18;border:1.5px solid #3AACAC;border-radius:8px;'
                'padding:9px 18px;margin:22px 0 6px 0;color:#3AACAC;font-weight:700;font-size:0.92rem;">'
                '🗂&nbsp; CARTE DU JOUR — Articles permanents</div>',
                unsafe_allow_html=True)
            for sub in CAT_MENU:
                items   = carte.get(sub,[])
                if not items: continue
                color   = CAT_COLORS[sub]
                icon    = CAT_ICONS.get(sub,"")
                cmd_key = f"Carte · {sub}"
                selected = existing.get(cmd_key,[])
                st.markdown(cat_header(color, f"↳ {sub.upper()} (carte)", icon),
                            unsafe_allow_html=True)
                st.markdown(
                    f'<div style="background:#22263D;border:1px solid {color}55;'
                    f'border-top:none;border-radius:0 0 8px 8px;padding:8px 4px 4px 4px;'
                    f'margin-bottom:10px;"></div>',
                    unsafe_allow_html=True)
                cols = st.columns(2)
                for i, item in enumerate(items):
                    with cols[i % 2]:
                        if st.checkbox(item, value=(item in selected),
                                       key=f"cb_{wk}_{sal}_{jour}_{cmd_key}_{i}"):
                            choix.setdefault(cmd_key,[]).append(item)

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("✓  Valider ma commande", type="primary", use_container_width=True):
            st.session_state.commandes.setdefault(wk,{}).setdefault(sal,{})[jour] = choix
            save_commandes_wk(wk, st.session_state.commandes[wk])
            total = sum(len(v) for v in choix.values())
            if total: st.success(f"✓ {total} article(s) enregistré(s) pour **{jour}**.")
            else:     st.warning(f"Commande effacée pour {jour}.")
            st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE : SAISIR LE MENU
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "menu":
    st.markdown("## ⚙ Menu de la semaine")
    wk   = st.session_state.week_sel
    menu = st.session_state.menus.get(wk, {
        "semaine":wk,"periode":"",
        "jours":{j:{c:[] for c in CAT_MENU} for j in JOURS}})

    st.markdown("### 📥 Importer le fichier du traiteur")
    uploaded = st.file_uploader("Fichier `.xls` reçu du traiteur",
                                 type=["xls","html","htm"], key="uploader_menu")
    if uploaded:
        try:
            parsed = parse_traiteur_html(uploaded.read())
            wk_imp = parsed["semaine"]
            st.info(f"Semaine détectée : **{week_label(wk_imp)}**" +
                    (" *(déjà existante — sera remplacée)*" if wk_imp in st.session_state.menus else ""))
            if st.button("✅ Confirmer l'import", type="primary"):
                st.session_state.menus[wk_imp] = parsed
                save_menu(wk_imp, parsed)
                st.session_state.week_sel = wk_imp
                st.success(f"Menu importé — {sum(len(v) for j in parsed['jours'].values() for v in j.values())} articles.")
                st.rerun()
        except ModuleNotFoundError:
            st.error("❌ `pip install beautifulsoup4`")
        except Exception as e:
            st.error(f"❌ {e}")

    st.divider()
    st.markdown("### ✏️ Saisie manuelle")

    # Même logique que la page admin : vérifier la cohérence de la période stockée
    periode_stockee  = clean_periode(menu.get("periode",""), wk)
    if _periode_coherente(periode_stockee, wk):
        periode_defaut = periode_stockee
    else:
        periode_defaut = week_label(wk)

    periode = st.text_input("Période (ex: Du 02 juin au 06 juin 2026)",
                             value=periode_defaut, key=f"periode_input_{wk}")
    tabs = st.tabs(JOURS); new_jours = {}
    for t, jour in zip(tabs, JOURS):
        with t:
            new_jours[jour] = {}
            jour_data = menu.get("jours",{}).get(jour,{})
            for cat in CAT_MENU:
                color = CAT_COLORS[cat]; icon = CAT_ICONS.get(cat,"")
                st.markdown(inline_label(color,icon,cat), unsafe_allow_html=True)
                val = "\n".join(jour_data.get(cat,[]))
                txt = st.text_area("Plats", value=val, height=110,
                                   key=f"menu_{wk}_{jour}_{cat}",
                                   label_visibility="collapsed")
                new_jours[jour][cat] = [l.strip() for l in txt.split("\n") if l.strip()]

    if st.button("💾 Enregistrer le menu", type="primary"):
        m = {"semaine":wk,"periode":periode,"jours":new_jours}
        st.session_state.menus[wk] = m
        save_menu(wk, m)
        st.success("Menu enregistré !"); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE : CARTE PERMANENTE
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "carte":
    st.markdown("## 🗂 Carte permanente")
    st.markdown("*Articles proposés tous les jours en complément du menu.*")
    carte = st.session_state.carte; new_carte = {}
    tabs = st.tabs(CAT_MENU)
    for t, cat in zip(tabs, CAT_MENU):
        with t:
            color = CAT_COLORS[cat]; icon = CAT_ICONS.get(cat,"")
            st.markdown(inline_label(color,icon,cat), unsafe_allow_html=True)
            val = "\n".join(carte.get(cat,[]))
            txt = st.text_area("Articles", value=val, height=200,
                               key=f"carte_{cat}", label_visibility="collapsed")
            new_carte[cat] = [l.strip() for l in txt.split("\n") if l.strip()]
    c1,c2 = st.columns([2,1])
    with c1:
        if st.button("💾 Enregistrer la carte", type="primary", use_container_width=True):
            st.session_state.carte = new_carte
            save_carte(new_carte)
            st.success(f"Carte enregistrée — {sum(len(v) for v in new_carte.values())} article(s).")
            st.rerun()
    with c2:
        if st.button("↺ Réinitialiser", use_container_width=True):
            st.session_state.carte = dict(DEFAULT_CARTE)
            save_carte(DEFAULT_CARTE)
            st.info("Réinitialisée."); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# PAGE : SALARIÉS
# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "salaries":
    st.markdown("## 👥 Gestion des salariés")
    salaries = list(st.session_state.salaries)
    for i, sal in enumerate(salaries):
        c1,c2 = st.columns([4,1])
        with c1:
            salaries[i] = st.text_input(f"Salarié {i+1}", value=sal,
                                         key=f"sal_edit_{i}", label_visibility="collapsed")
        with c2:
            if st.button("✕", key=f"sal_del_{i}"):
                salaries.pop(i)
                st.session_state.salaries = salaries
                save_salaries(salaries); st.rerun()
    st.divider()
    new_sal = st.text_input("➕ Nouveau salarié", placeholder="Prénom NOM",
                             key="new_sal_input")
    if st.button("Ajouter", type="primary") and new_sal.strip():
        salaries.append(new_sal.strip())
        st.session_state.salaries = salaries
        save_salaries(salaries); st.rerun()
    if st.button("💾 Enregistrer", use_container_width=True):
        st.session_state.salaries = [s for s in salaries if s.strip()]
        save_salaries(st.session_state.salaries)
        st.success("Liste mise à jour."); st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "admin":
    st.markdown("## 📊 Exports & Administration")
    wk        = st.session_state.week_sel
    menu      = st.session_state.menus.get(wk,{})
    commandes = st.session_state.commandes.get(wk,{})
    salaries  = st.session_state.salaries
    carte     = st.session_state.carte

    # Période : on nettoie ce qui est stocké, MAIS si la période nettoyée
    # ne correspond pas à la semaine sélectionnée, on force week_label(wk)
    periode_stockee = clean_periode(menu.get("periode",""), wk)

    if _periode_coherente(periode_stockee, wk):
        periode = periode_stockee
    else:
        periode = week_label(wk)

    st.markdown(f"**Semaine :** {week_label(wk)}  —  *{periode}*")

    st.markdown("### 📤 Bon de commande traiteur")
    buf = build_export_traiteur(menu, commandes, salaries, carte, periode, wk)
    st.download_button("⬇ Télécharger (Excel)", data=buf,
                       file_name=f"BonCommande_Traiteur_{wk}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary", use_container_width=True)

    st.markdown("---")
    st.markdown("### 📋 Récapitulatif interne")
    buf2 = build_export_interne(menu, commandes, salaries, carte, periode, wk)
    st.download_button("⬇ Télécharger (Excel)", data=buf2,
                       file_name=f"Recapitulatif_Interne_{wk}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)

    st.markdown("---")
    st.markdown("### 💰 Export déduction paie")
    st.markdown("*Repas complet vs plat unique, par salarié et par jour, sur une ou plusieurs semaines.*")

    # Sélection de la plage de semaines
    all_wks   = weeks_list(st.session_state.menus)
    all_labels = [week_label(k) for k in all_wks]

    col_from, col_to = st.columns(2)
    with col_from:
        idx_from = st.selectbox("Semaine de début", all_labels,
                                 index=len(all_labels)-1,
                                 key="paie_from")
    with col_to:
        idx_to = st.selectbox("Semaine de fin", all_labels,
                               index=0,
                               key="paie_to")

    wk_from = all_wks[all_labels.index(idx_from)]
    wk_to   = all_wks[all_labels.index(idx_to)]

    # Trier pour que from <= to
    if wk_from > wk_to:
        wk_from, wk_to = wk_to, wk_from

    semaines_sel = [(k, week_label(k),
                     (lambda p, k: p if _periode_coherente(p, k) else week_label(k))(
                         clean_periode(st.session_state.menus.get(k,{}).get("periode",""), k), k))
                    for k in all_wks if wk_from <= k <= wk_to]

    st.caption(f"{len(semaines_sel)} semaine(s) sélectionnée(s) : "
               f"{', '.join(s[1] for s in semaines_sel)}")

    # Légende
    col_leg1, col_leg2 = st.columns(2)
    with col_leg1:
        st.markdown("🟢 **Repas complet** = plat + entrée et/ou produit laitier et/ou dessert")
    with col_leg2:
        st.markdown("🟡 **Plat unique** = plats garnis / accompagnements uniquement")

    if semaines_sel:
        # Aperçu rapide
        with st.expander("👁 Aperçu des repas pour la semaine sélectionnée"):
            preview_cmds = st.session_state.commandes.get(wk, {})
            prev_rows = []
            for sal in salaries:
                row_p = {"Salarié": sal}
                for jour in JOURS:
                    t = classifier_repas(preview_cmds.get(sal,{}).get(jour,{}))
                    row_p[jour] = {"complet":"🟢 Complet","plat_unique":"🟡 Plat","rien":"—"}[t]
                prev_rows.append(row_p)
            st.table(prev_rows)

        buf_paie = build_export_paie(
            st.session_state.commandes, salaries, semaines_sel)
        fname = (f"DeductionPaie_{wk_from}_au_{wk_to}.xlsx"
                 if wk_from != wk_to else f"DeductionPaie_{wk_from}.xlsx")
        st.download_button(
            "⬇ Télécharger export déduction paie (Excel)",
            data=buf_paie,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    else:
        st.warning("Aucune semaine dans la plage sélectionnée.")
    if not commandes:
        st.info("Aucune commande enregistrée pour cette semaine.")
    else:
        recap = []
        for sal in salaries:
            row = {"Salarié": sal}
            for jour in JOURS:
                total = sum(len(v) for v in commandes.get(sal,{}).get(jour,{}).values())
                row[jour] = f"✓ {total}" if total else "—"
            recap.append(row)
        st.table(recap)

    st.markdown("---")
    st.markdown("### 🗓 Nouvelle semaine")
    c1,c2 = st.columns(2)
    with c1:
        new_wk = st.text_input("Clé semaine", placeholder="ex: 2026-S30")
        if st.button("Créer", type="primary") and new_wk.strip():
            nwk = new_wk.strip()
            if re.match(r'\d{4}-S\d{2}', nwk):
                if nwk not in st.session_state.menus:
                    m = {"semaine":nwk,"periode":"",
                         "jours":{j:{c:[] for c in CAT_MENU} for j in JOURS}}
                    st.session_state.menus[nwk] = m
                    save_menu(nwk, m)
                st.session_state.week_sel = nwk
                st.success(f"Semaine {week_label(nwk)} créée."); st.rerun()
            else: st.error("Format : 2026-S30")
    with c2:
        st.markdown("**Semaines disponibles :**")
        for k in weeks_list(st.session_state.menus):
            n = sum(sum(len(v) for v in commandes.get(k,{}).get(s,{}).values())
                    for s in salaries)
            st.markdown(f"- `{k}` — {week_label(k)} — {n} article(s)")

    st.markdown("---")
    st.markdown("### 🔄 Recharger les données")
    if st.button("🔄 Recharger", use_container_width=True):
        del st.session_state["loaded"]
        if _use_gsheets():
            _gs_client.clear()
        st.rerun()

    # Diagnostic Google Sheets si configuré
    if _use_gsheets():
        st.markdown("---")
        st.markdown("### 🔍 Diagnostic Google Sheets")
        if st.button("🔍 Tester la connexion"):
            try:
                sh   = _gs_client().open_by_key(st.secrets["gsheet"]["spreadsheet_id"])
                tabs = [ws.title for ws in sh.worksheets()]
                st.success(f"✅ Connexion OK — Sheet : **{sh.title}** — Onglets : {tabs}")
            except Exception as e:
                st.error(f"❌ {e}")
        if st.button("📝 Écriture test"):
            try:
                _gs_save("config","_test",{"ok":True,"ts":str(datetime.now())})
                st.success("✅ Écriture réussie dans l'onglet `config` !")
            except Exception as e:
                st.error(f"❌ {e}")
